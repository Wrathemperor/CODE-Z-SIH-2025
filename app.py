import os
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, flash, redirect, url_for, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler, LabelEncoder
from datetime import date, datetime
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from openpyxl.styles import PatternFill


app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = 'a-very-secret-key-that-should-be-changed'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your_email@gmail.com'
app.config['MAIL_PASSWORD'] = 'your_app_password'


db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(150), nullable=False)
    students = db.relationship('StudentData', backref='uploader', lazy=True, cascade="all, delete-orphan")

class StudentData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    internal_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    student_id_str = db.Column(db.String(100), name="Student ID")
    student_name = db.Column(db.String(100), name="Student Name")
    parent_mail = db.Column(db.String(100), name="Parent Mail")
    raw_data = db.Column(db.JSON)
    dropout_probability = db.Column(db.Float)
    risk_level = db.Column(db.String(50))
    remarks = db.Column(db.String(500))
    attendances = db.relationship('Attendance', backref='student', lazy=True, cascade="all, delete-orphan")
    counselling_sessions = db.relationship('Counselling', backref='student', lazy=True, cascade="all, delete-orphan")

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_data_id = db.Column(db.Integer, db.ForeignKey('student_data.id'), nullable=False)
    week_str = db.Column(db.String(20), nullable=False)
    subject_attendance = db.Column(db.JSON, nullable=False)

class Counselling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_data_id = db.Column(db.Integer, db.ForeignKey('student_data.id'), nullable=False)
    counselling_date = db.Column(db.String(20), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

model_2sem, scaler_2sem, label_encoders_2sem, model_columns_2sem = None, None, {}, None
model_1sem, scaler_1sem, label_encoders_1sem, model_columns_1sem = None, None, {}, None
REQUIRED_COLUMNS_1SEM = [ "Marital status", "Course", "Daytime/evening attendance", "Previous qualification", "Previous qualification (grade)", "Mother's qualification", "Father's qualification", "Mother's occupation", "Father's occupation", "Displaced", "Educational special needs", "Debtor", "Tuition fees up to date", "Gender", "Scholarship holder", "Age at enrollment", "International", "Curricular units 1st sem (credited)", "Curricular units 1st sem (enrolled)", "Curricular units 1st sem (evaluations)", "Curricular units 1st sem (approved)", "Curricular units 1st sem (grade)", "Curricular units 1st sem (without evaluations)" ]
REQUIRED_COLUMNS_2SEM = REQUIRED_COLUMNS_1SEM + [ "Curricular units 2nd sem (credited)", "Curricular units 2nd sem (enrolled)", "Curricular units 2nd sem (evaluations)", "Curricular units 2nd sem (approved)", "Curricular units 2nd sem (grade)", "Curricular units 2nd sem (without evaluations)" ]
COURSE_MAP = {33: "Biofuel Production Technologies", 171: "Animation and Multimedia Design", 8014: "Social Service (evening)", 9003: "Agronomy", 9070: "Communication Design", 9085: "Veterinary Nursing", 9119: "Informatics Engineering", 9130: "Equinculture", 9147: "Management", 9238: "Social Service", 9254: "Tourism", 9500: "Nursing", 9556: "Oral Hygiene", 9670: "Advertising and Marketing Management", 9773: "Journalism and Communication", 9853: "Basic Education", 9991: "Management (evening)"}

def _preprocess_data(df, is_training=True, encoders=None, fit_scaler=None, model_cols=None):
    processed_df = df.copy()
    if is_training:
        processed_df['Target'] = processed_df['Target'].apply(lambda x: 1 if x == 'Dropout' else 0)
        label_encoders = {}
        for col in sorted(processed_df.select_dtypes(include=['object']).columns):
            if col != 'Target':
                le = LabelEncoder()
                processed_df[col] = le.fit_transform(processed_df[col].astype(str))
                label_encoders[col] = le
        X = processed_df.drop('Target', axis=1)
        y = processed_df['Target']
        model_columns = X.columns.tolist()
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        return X_scaled, y, scaler, label_encoders, model_columns
    else:
        for col in model_cols:
            if col not in processed_df.columns: processed_df[col] = 0
        processed_df = processed_df[model_cols]
        for col, le in encoders.items():
            if col in processed_df.columns:
                known_labels = list(le.classes_)
                processed_df[col] = processed_df[col].astype(str).apply(lambda x: x if x in known_labels else '<unknown>')
                if '<unknown>' not in le.classes_: le.classes_ = np.append(le.classes_, '<unknown>')
                processed_df[col] = le.transform(processed_df[col])
        return fit_scaler.transform(processed_df)

def train_svm_model_on_startup():
    global model_1sem, scaler_1sem, label_encoders_1sem, model_columns_1sem, model_2sem, scaler_2sem, label_encoders_2sem, model_columns_2sem
    id_cols = ['Student ID', 'Student Name', 'Parent Mail']
    try:
        df_1sem_raw = pd.read_csv('Data(Nosem).csv')
        df_1sem_raw.rename(columns={'Daytime/evening attendance\t': 'Daytime/evening attendance', 'Student Mail': 'Parent Mail'}, inplace=True)
        X_train, y_train, scaler, encoders, model_cols = _preprocess_data(df_1sem_raw.drop(columns=id_cols, errors='ignore'), is_training=True)
        scaler_1sem, label_encoders_1sem, model_columns_1sem = scaler, encoders, model_cols
        model_1sem = SVC(kernel='rbf', probability=True, class_weight='balanced', random_state=42).fit(X_train, y_train)
        print("✅ 1-Semester SVM model trained successfully.")
    except Exception as e: print(f"❌ Error training 1-Semester model: {e}")
    try:
        df_2sem_raw = pd.read_csv('Data(Sem-2).csv')
        df_2sem_raw.rename(columns={'Daytime/evening attendance\t': 'Daytime/evening attendance', 'Student Mail': 'Parent Mail'}, inplace=True)
        X_train, y_train, scaler, encoders, model_cols = _preprocess_data(df_2sem_raw.drop(columns=id_cols, errors='ignore'), is_training=True)
        scaler_2sem, label_encoders_2sem, model_columns_2sem = scaler, encoders, model_cols
        model_2sem = SVC(kernel='rbf', probability=True, class_weight='balanced', random_state=42).fit(X_train, y_train)
        print("✅ 2-Semester SVM model trained successfully.")
    except Exception as e: print(f"❌ Error training 2-Semester model: {e}")

def generate_remarks(row):
    remarks = []
    if row.get('Tuition fees up to date') == 0: remarks.append("Tuition fees not up to date")
    if row.get('Previous qualification (grade)', 200) < 110: remarks.append("Low previous grade")
    return ", ".join(remarks[:2]) if remarks else "Multiple contributing factors."

def send_email(parent_email, subject, body):
    if app.config['MAIL_USERNAME'] == 'your_email@gmail.com':
        print("⚠️ Email not sent: MAIL_USERNAME not configured.")
        return False, "Email not configured on server."
    try:
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = parent_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        server.send_message(msg)
        server.quit()
        return True, "Email sent successfully."
    except Exception as e:
        print(f"❌ Failed to send email: {e}")
        return False, str(e)

def create_excel_report(students):
    output = io.BytesIO()
    data_to_excel = []
    for s in students:
        row = s.raw_data.copy()
        row['Risk_Level'] = s.risk_level
        row['Dropout_Probability'] = s.dropout_probability
        row['Remarks'] = s.remarks
        data_to_excel.append(row)

    df = pd.DataFrame(data_to_excel)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Predictions')
        worksheet = writer.sheets['Predictions']
        fills = {'High': PatternFill(start_color='FFC7CE', fill_type='solid'), 'Medium': PatternFill(start_color='FFEB9C', fill_type='solid'), 'Low': PatternFill(start_color='C6EFCE', fill_type='solid')}
        
        if 'Risk_Level' in df.columns:
            risk_col_idx = df.columns.get_loc('Risk_Level') + 1
            for row_idx, student in enumerate(students, start=2):
                risk_val = student.risk_level
                if risk_val in fills:
                    for cell in worksheet[row_idx]:
                        cell.fill = fills[risk_val]
    output.seek(0)
    return output

def process_student_data(student):
    raw = student.raw_data
    return {
        'id': student.id,
        'internal_id': student.internal_id,
        'summary': {
            'Student Name': student.student_name,
            'Student ID': student.student_id_str,
            'Parent Mail': student.parent_mail,
            'Course': COURSE_MAP.get(raw.get('Course'), 'N/A'),
            'Risk_Level': student.risk_level,
            'Dropout_Probability': student.dropout_probability,
            'Remarks': student.remarks,
        },
        'academic': [
            {'label_en': 'Previous Qualification', 'label_hi': 'पिछली योग्यता', 'value': raw.get('Previous qualification', 'N/A')},
            {'label_en': 'Previous Qualification Grade', 'label_hi': 'पिछली योग्यता ग्रेड', 'value': raw.get('Previous qualification (grade)', 'N/A')},
            {'label_en': '1st Sem Approved Units', 'label_hi': 'प्रथम सेम स्वीकृत इकाइयां', 'value': raw.get('Curricular units 1st sem (approved)', 'N/A')},
            {'label_en': '1st Sem Grade', 'label_hi': 'प्रथम सेम ग्रेड', 'value': raw.get('Curricular units 1st sem (grade)', 'N/A')},
            {'label_en': '2nd Sem Approved Units', 'label_hi': 'द्वितीय सेम स्वीकृत इकाइयां', 'value': raw.get('Curricular units 2nd sem (approved)', 'N/A')},
            {'label_en': '2nd Sem Grade', 'label_hi': 'द्वितीय सेम ग्रेड', 'value': raw.get('Curricular units 2nd sem (grade)', 'N/A')},
        ],
        'demographics': [
            {'label_en': 'Gender', 'label_hi': 'लिंग', 'value': 'Male' if raw.get('Gender') == 1 else 'Female'},
            {'label_en': 'Age at Enrollment', 'label_hi': 'नामांकन के समय आयु', 'value': raw.get('Age at enrollment', 'N/A')},
            {'label_en': 'Marital Status', 'label_hi': 'वैवाहिक स्थिति', 'value': raw.get('Marital status', 'N/A')},
            {'label_en': 'International Student', 'label_hi': 'अंतर्राष्ट्रीय छात्र', 'value': 'Yes' if raw.get('International') == 1 else 'No'},
            {'label_en': 'Displaced', 'label_hi': 'विस्थापित', 'value': 'Yes' if raw.get('Displaced') == 1 else 'No'},
            {'label_en': 'Special Needs', 'label_hi': 'विशेष आवश्यकताएं', 'value': 'Yes' if raw.get('Educational special needs') == 1 else 'No'},
        ],
        'financial': [
            {'label_en': 'Tuition Fees Up to Date', 'label_hi': 'ट्यूशन फीस अद्यतित', 'value': 'Yes' if raw.get('Tuition fees up to date') == 1 else 'No'},
            {'label_en': 'Debtor', 'label_hi': 'देनदार', 'value': 'Yes' if raw.get('Debtor') == 1 else 'No'},
            {'label_en': 'Scholarship Holder', 'label_hi': 'छात्रवृत्ति धारक', 'value': 'Yes' if raw.get('Scholarship holder') == 1 else 'No'},
        ],
        'family': [
            {'label_en': "Mother's Qualification", 'label_hi': 'माता की योग्यता', 'value': raw.get("Mother's qualification", 'N/A')},
            {'label_en': "Father's Qualification", 'label_hi': 'पिता की योग्यता', 'value': raw.get("Father's qualification", 'N/A')},
            {'label_en': "Mother's Occupation", 'label_hi': 'माता का व्यवसाय', 'value': raw.get("Mother's occupation", 'N/A')},
            {'label_en': "Father's Occupation", 'label_hi': 'पिता का व्यवसाय', 'value': raw.get("Father's occupation", 'N/A')},
        ]
    }

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('index'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user and check_password_hash(user.password_hash, request.form.get('password')):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated: return redirect(url_for('index'))
    if request.method == 'POST':
        if User.query.filter_by(username=request.form.get('username')).first():
            flash('Username already exists.', 'danger')
        else:
            hashed_pw = generate_password_hash(request.form.get('password'), method='pbkdf2:sha256')
            new_user = User(username=request.form.get('username'), password_hash=hashed_pw)
            db.session.add(new_user)
            db.session.commit()
            flash('Account created! Please log in.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    data_type = request.form.get('data_type', '2sem')
    model, scaler, encoders, req_cols, model_cols = (model_2sem, scaler_2sem, label_encoders_2sem, REQUIRED_COLUMNS_2SEM, model_columns_2sem) if data_type == '2sem' else (model_1sem, scaler_1sem, label_encoders_1sem, REQUIRED_COLUMNS_1SEM, model_columns_1sem)
    file = request.files.get('file')
    if not file:
        flash('No file selected.', 'warning')
        return redirect(url_for('index'))

    try:
        file_contents = io.BytesIO(file.read())
        df_raw = pd.read_excel(file_contents) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file_contents)
        df_raw.rename(columns={'Daytime/evening attendance\t': 'Daytime/evening attendance', 'Student Mail': 'Parent Mail'}, inplace=True)
        
        X_predict_scaled = _preprocess_data(df_raw.copy(), is_training=False, encoders=encoders, fit_scaler=scaler, model_cols=model_cols)
        probabilities = model.predict_proba(X_predict_scaled)[:, 1]
        
        results_df = df_raw.copy()
        results_df['Dropout_Probability'] = probabilities
        results_df['Risk_Level'] = pd.cut(probabilities, bins=[-0.1, 0.4, 0.7, 1.1], labels=['Low', 'Medium', 'High'], right=False).astype(str)
        results_df['Remarks'] = results_df.apply(generate_remarks, axis=1)

        for index, row in results_df.iterrows():
            row_dict = {k: (int(v) if isinstance(v, np.integer) else float(v) if isinstance(v, np.floating) else v) for k, v in row.to_dict().items()}
            new_student = StudentData(
                internal_id=index, user_id=current_user.id,
                student_id_str=str(row.get('Student ID', 'N/A')),
                student_name=row.get('Student Name', 'N/A'),
                parent_mail=row.get('Parent Mail', 'N/A'),
                raw_data=row_dict, dropout_probability=row['Dropout_Probability'],
                risk_level=row['Risk_Level'], remarks=row['Remarks']
            )
            db.session.add(new_student)
        db.session.commit()
        return redirect(url_for('dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred: {e}', 'danger')
        return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    students_exist = StudentData.query.filter_by(user_id=current_user.id).first()
    if not students_exist:
        flash("No student data found. Please upload a file to begin.", "info")
        return redirect(url_for('index'))
    return render_template('dashboard.html')

@app.route('/student/<int:student_db_id>')
@login_required
def student_detail(student_db_id):
    student = db.session.get(StudentData, student_db_id)
    if student is None or student.user_id != current_user.id:
        flash('Student not found.', 'danger')
        return redirect(url_for('dashboard'))
    
    processed_student = process_student_data(student)
    return render_template('student_detail.html', student=processed_student, internal_id=student.id)

@app.route('/counselling')
@login_required
def counselling():
    day = request.args.get('day', date.today().isoformat())
    sessions = Counselling.query.join(StudentData).filter(
        StudentData.user_id == current_user.id,
        Counselling.counselling_date == day
    ).all()
    
    students = [process_student_data(session.student) for session in sessions]
    return render_template('counselling.html', students=students, day=day)

@app.route('/upload_attendance', methods=['POST'])
@login_required
def upload_attendance():
    week_str = request.form.get('week')
    if not week_str:
        flash('Week not selected.', 'danger')
        return redirect(url_for('dashboard'))

    file = request.files.get('file')
    if not file:
        flash('No attendance file selected.', 'warning')
        return redirect(url_for('dashboard'))

    try:
        file_contents = io.BytesIO(file.read())
        df_attendance = pd.read_excel(file_contents) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file_contents)

        required_col = 'Student_ID'
        df_attendance.rename(columns={'Student ID': 'Student_ID'}, inplace=True)


        if required_col not in df_attendance.columns:
            flash(f"Attendance file must contain a '{required_col}' column.", 'danger')
            return redirect(url_for('dashboard'))

        subjects = [col for col in df_attendance.columns if col != required_col]
        
        updated_count = 0
        not_found_students = []

        for index, row in df_attendance.iterrows():
            student_id_str = str(row[required_col])
            
            student = StudentData.query.filter_by(user_id=current_user.id, student_id_str=student_id_str).first()

            if student:
                attendance_values = {subject: float(row[subject]) for subject in subjects}

                att_record = Attendance.query.filter_by(student_data_id=student.id, week_str=week_str).first()

                if not att_record:
                    att_record = Attendance(student_data_id=student.id, week_str=week_str)
                    db.session.add(att_record)
                
                att_record.subject_attendance = attendance_values
                updated_count += 1
            else:
                not_found_students.append(student_id_str)

        db.session.commit()
        
        flash(f'Successfully updated attendance for {updated_count} students.', 'success')
        if not_found_students:
            flash(f"Could not find the following Student IDs: {', '.join(not_found_students)}", 'warning')

        return redirect(url_for('dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while processing the attendance file: {e}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/api/assign-counselling', methods=['POST'])
@login_required
def assign_counselling():
    data = request.json
    student_db_id = data.get('student_db_id')
    day = data.get('day')
    
    existing_session = Counselling.query.filter_by(student_data_id=student_db_id, counselling_date=day).first()
    if existing_session:
        return jsonify({'success': False, 'message': 'Student already assigned for this day.'})

    new_session = Counselling(student_data_id=student_db_id, counselling_date=day)
    db.session.add(new_session)
    db.session.commit()
    
    student = db.session.get(StudentData, student_db_id)
    if student and student.parent_mail:
        subject = f"Counselling Session for {student.student_name}"
        body = f"Dear Parent,\n\nA counselling session for {student.student_name} has been scheduled for {day}.\n\nSincerely,\nThe School Administration"
        success, msg = send_email(student.parent_mail, subject, body)
        return jsonify({'success': True, 'message': f'Student assigned for {day}. {msg}'.strip()})
    
    return jsonify({'success': True, 'message': f'Student assigned for {day}. Email not sent (no address).'})

@app.route('/api/end-counselling', methods=['POST'])
@login_required
def end_counselling():
    data = request.json
    day = data.get('day')
    student_db_id = data.get('student_db_id')
    
    session = Counselling.query.join(StudentData).filter(
        StudentData.user_id == current_user.id,
        Counselling.student_data_id == student_db_id,
        Counselling.counselling_date == day
    ).first()

    if session:
        db.session.delete(session)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Session ended.'})
    return jsonify({'success': False, 'message': 'Session not found.'})

@app.route('/api/reschedule-counselling', methods=['POST'])
@login_required
def reschedule_counselling():
    data = request.json
    old_day = data.get('old_day')
    new_day = data.get('new_day')
    student_db_id = data.get('student_db_id')

    session = Counselling.query.join(StudentData).filter(
        StudentData.user_id == current_user.id,
        Counselling.student_data_id == student_db_id,
        Counselling.counselling_date == old_day
    ).first()

    if not session:
        return jsonify({'success': False, 'message': 'Original session not found.'})

    existing_new = Counselling.query.filter_by(student_data_id=student_db_id, counselling_date=new_day).first()
    if existing_new:
        return jsonify({'success': False, 'message': 'Student already has a session on the new date.'})

    session.counselling_date = new_day
    db.session.commit()
    
    if session.student and session.student.parent_mail:
        subject = f"RESCHEDULED: Counselling for {session.student.student_name}"
        body = f"Dear Parent,\n\nPlease note that the counselling session for {session.student.student_name} has been rescheduled to {new_day}.\n\nSincerely,\nThe School Administration"
        send_email(session.student.parent_mail, subject, body)

    return jsonify({'success': True, 'message': f'Rescheduled to {new_day}.'})

@app.route('/api/dashboard-data')
@login_required
def get_dashboard_data():
    students = StudentData.query.filter_by(user_id=current_user.id).all()
    if not students:
        return jsonify({'summary': {}, 'students': []})

    summary = {
        'total_students': len(students),
        'high_risk': sum(1 for s in students if s.risk_level == 'High'),
        'medium_risk': sum(1 for s in students if s.risk_level == 'Medium'),
        'low_risk': sum(1 for s in students if s.risk_level == 'Low')
    }
    
    student_list = [{
        'internal_id': s.id,
        'Student ID': s.student_id_str,
        'Student Name': s.student_name,
        'Course': COURSE_MAP.get(s.raw_data.get('Course'), 'Unknown'),
        'Risk_Level': s.risk_level,
        'Dropout_Probability': s.dropout_probability
    } for s in students]
    
    student_list.sort(key=lambda x: x['Dropout_Probability'], reverse=True)
    return jsonify({'summary': summary, 'students': student_list})

@app.route('/api/chart-data')
@login_required
def get_chart_data():
    students = StudentData.query.filter_by(user_id=current_user.id).all()
    
    course_risk = {}
    for s in students:
        course_name = COURSE_MAP.get(s.raw_data.get('Course'), 'Unknown')
        if course_name not in course_risk:
            course_risk[course_name] = {'High': 0, 'Medium': 0, 'Low': 0}
        course_risk[course_name][s.risk_level] += 1
        
    course_labels = list(course_risk.keys())
    high_risk_counts = [course_risk[c]['High'] for c in course_labels]
    medium_risk_counts = [course_risk[c]['Medium'] for c in course_labels]
    low_risk_counts = [course_risk[c]['Low'] for c in course_labels]
    
    high_risk_students = [s for s in students if s.risk_level == 'High']
    sem1_approved_sum = sum(s.raw_data.get('Curricular units 1st sem (approved)', 0) for s in high_risk_students)
    sem2_approved_sum = sum(s.raw_data.get('Curricular units 2nd sem (approved)', 0) for s in high_risk_students)
    num_high_risk = len(high_risk_students) if high_risk_students else 1
    
    sem_perf_data = {
        'labels': ['1st Semester', '2nd Semester'],
        'datasets': [{'label': 'Average Approved Units', 'data': [sem1_approved_sum / num_high_risk, sem2_approved_sum / num_high_risk], 'backgroundColor': ['rgba(54, 162, 235, 0.6)', 'rgba(255, 159, 64, 0.6)'], 'borderColor': ['rgba(54, 162, 235, 1)', 'rgba(255, 159, 64, 1)'], 'borderWidth': 1}]
    }

    return jsonify({
        'course_vs_risk': {
            'labels': course_labels,
            'datasets': [
                {'label': 'High Risk', 'data': high_risk_counts, 'backgroundColor': 'rgba(239, 68, 68, 0.7)'},
                {'label': 'Medium Risk', 'data': medium_risk_counts, 'backgroundColor': 'rgba(245, 158, 11, 0.7)'},
                {'label': 'Low Risk', 'data': low_risk_counts, 'backgroundColor': 'rgba(34, 197, 94, 0.7)'}
            ]
        },
        'semester_performance_high_risk': sem_perf_data
    })


@app.route('/download_excel')
@login_required
def download_excel():
    students = StudentData.query.filter_by(user_id=current_user.id).all()
    if not students:
        flash('No data to download.', 'warning'); return redirect(url_for('dashboard'))
    excel_output = create_excel_report(students)
    return send_file(excel_output, as_attachment=True, download_name='prediction_results.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/sample-data')
@login_required
def sample_data():
    return send_file('sample_student_data.csv', as_attachment=True)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='teacher').first():
            db.session.add(User(username='teacher', password_hash=generate_password_hash('password', method='pbkdf2:sha256')))
            db.session.commit()
    train_svm_model_on_startup()
    app.run(debug=True, host='0.0.0.0', port=5000)